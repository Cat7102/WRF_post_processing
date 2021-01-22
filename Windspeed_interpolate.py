# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Windspeed_interpolate.ui'
#
# Created by: PyQt5 UI code generator 5.15.2
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.


from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import QThread, pyqtSignal
import sys
import netCDF4 as nc
from wrf import getvar,to_np,ALL_TIMES
from datetime import datetime,timedelta
from scipy.interpolate.rbf import Rbf
import openpyxl
import numpy as np


class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(445, 238)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(1)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(Dialog.sizePolicy().hasHeightForWidth())
        Dialog.setSizePolicy(sizePolicy)
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(Dialog)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label = QtWidgets.QLabel(Dialog)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        self.lineEdit = QtWidgets.QLineEdit(Dialog)
        self.lineEdit.setObjectName("lineEdit")
        self.horizontalLayout.addWidget(self.lineEdit)
        self.pushButton_2 = QtWidgets.QPushButton(Dialog)
        self.pushButton_2.setObjectName("pushButton_2")
        self.horizontalLayout.addWidget(self.pushButton_2)
        self.verticalLayout_2.addLayout(self.horizontalLayout)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.label_2 = QtWidgets.QLabel(Dialog)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_2.sizePolicy().hasHeightForWidth())
        self.label_2.setSizePolicy(sizePolicy)
        self.label_2.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_2.addWidget(self.label_2)
        self.comboBox = QtWidgets.QComboBox(Dialog)
        self.comboBox.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.comboBox.sizePolicy().hasHeightForWidth())
        self.comboBox.setSizePolicy(sizePolicy)
        self.comboBox.setIconSize(QtCore.QSize(16, 16))
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.horizontalLayout_2.addWidget(self.comboBox)
        self.label_6 = QtWidgets.QLabel(Dialog)
        self.label_6.setObjectName("label_6")
        self.horizontalLayout_2.addWidget(self.label_6)
        self.comboBox_3 = QtWidgets.QComboBox(Dialog)
        self.comboBox_3.setObjectName("comboBox_3")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.horizontalLayout_2.addWidget(self.comboBox_3)
        self.horizontalLayout_2.setStretch(1, 1)
        self.verticalLayout_2.addLayout(self.horizontalLayout_2)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label_3 = QtWidgets.QLabel(Dialog)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout_3.addWidget(self.label_3)
        self.lineEdit_2 = QtWidgets.QLineEdit(Dialog)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_2.sizePolicy().hasHeightForWidth())
        self.lineEdit_2.setSizePolicy(sizePolicy)
        self.lineEdit_2.setMaximumSize(QtCore.QSize(70, 16777215))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.horizontalLayout_3.addWidget(self.lineEdit_2)
        self.label_4 = QtWidgets.QLabel(Dialog)
        self.label_4.setObjectName("label_4")
        self.horizontalLayout_3.addWidget(self.label_4)
        self.lineEdit_3 = QtWidgets.QLineEdit(Dialog)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_3.sizePolicy().hasHeightForWidth())
        self.lineEdit_3.setSizePolicy(sizePolicy)
        self.lineEdit_3.setMaximumSize(QtCore.QSize(70, 16777215))
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.horizontalLayout_3.addWidget(self.lineEdit_3)
        self.label_5 = QtWidgets.QLabel(Dialog)
        self.label_5.setObjectName("label_5")
        self.horizontalLayout_3.addWidget(self.label_5)
        self.comboBox_2 = QtWidgets.QComboBox(Dialog)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.comboBox_2.sizePolicy().hasHeightForWidth())
        self.comboBox_2.setSizePolicy(sizePolicy)
        self.comboBox_2.setMaximumSize(QtCore.QSize(16777215, 16777215))
        self.comboBox_2.setObjectName("comboBox_2")
        self.horizontalLayout_3.addWidget(self.comboBox_2)
        self.verticalLayout_2.addLayout(self.horizontalLayout_3)
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.pushButton = QtWidgets.QPushButton(Dialog)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton.sizePolicy().hasHeightForWidth())
        self.pushButton.setSizePolicy(sizePolicy)
        self.pushButton.setIconSize(QtCore.QSize(16, 16))
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout_4.addWidget(self.pushButton)
        self.verticalLayout_2.addLayout(self.horizontalLayout_4)
        self.progressBar = QtWidgets.QProgressBar(Dialog)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.verticalLayout_2.addWidget(self.progressBar)

        self.retranslateUi(Dialog)
        self.comboBox_2.setCurrentIndex(-1)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "风速插值（高度）"))
        self.label.setText(_translate("Dialog", "图片路径："))
        self.pushButton_2.setText(_translate("Dialog", "选择"))
        self.label_2.setText(_translate("Dialog", "插值函数："))
        self.comboBox.setItemText(0, _translate("Dialog", "gaussian"))
        self.comboBox.setItemText(1, _translate("Dialog", "multiquadric"))
        self.comboBox.setItemText(2, _translate("Dialog", "inverse"))
        self.comboBox.setItemText(3, _translate("Dialog", "linear"))
        self.comboBox.setItemText(4, _translate("Dialog", "cubic"))
        self.comboBox.setItemText(5, _translate("Dialog", "quintic"))
        self.comboBox.setItemText(6, _translate("Dialog", "thin_plate"))
        self.label_6.setText(_translate("Dialog", "加速计算倍数（1x表示无加速）："))
        self.comboBox_3.setItemText(0, _translate("Dialog", "1x"))
        self.comboBox_3.setItemText(1, _translate("Dialog", "2x"))
        self.comboBox_3.setItemText(2, _translate("Dialog", "3x"))
        self.comboBox_3.setItemText(3, _translate("Dialog", "4x"))
        self.comboBox_3.setItemText(4, _translate("Dialog", "5x"))
        self.label_3.setText(_translate("Dialog", "经度："))
        self.label_4.setText(_translate("Dialog", "纬度："))
        self.label_5.setText(_translate("Dialog", "时间："))
        self.pushButton.setText(_translate("Dialog", "计算"))

        self.pushButton_2.clicked.connect(self.choosencfile)
        self.pushButton.clicked.connect(self.start_interpolate)
        self.thread=None #线程初始化

    def choosencfile(self):
        self.path, self.filetype = QtWidgets.QFileDialog.getOpenFileName(directory="C:/",filter="NC file (*.nc);;")
        self.lineEdit.setText(self.path)
        self.comboBox_2.addItems(self.get_ncfile_time(self.path))

    def get_ncfile_time(self,path):
        timelist=[]
        self.ncfile=nc.Dataset(path)
        time = getvar(self.ncfile, 'times')
        time = to_np(time)
        time = str(time)
        time = time[0:-10]
        times = getvar(self.ncfile, 'xtimes', timeidx=ALL_TIMES)
        formal_datetime = datetime.strptime(time, '%Y-%m-%dT%H:%M:%S')
        for i in times:
            timelist.append(str(formal_datetime + timedelta(minutes=int(i))))
        timelist.append('ALL')
        self.time_num=len(timelist)
        return timelist

    def start_interpolate(self):
        self.thread=WorkThread(self.path,self.lineEdit_2.text(),self.lineEdit_3.text(),self.comboBox.currentText(),
                               self.comboBox_2.currentIndex(),self.comboBox_2.currentText(),self.time_num,self.comboBox_3.currentIndex())
        self.thread.trigger.connect(self.call_back)
        self.thread.start()

    def call_back(self,progress):
        self.progressBar.setValue(progress)

    def write_to_xlsx(self):
        print()

class WorkThread(QThread):
    trigger = pyqtSignal(float)

    def __init__(self,path,input_lon,input_lat,function_type,time_num,time,all_time_index,speedup_index):
        # 初始化函数
        super(WorkThread, self).__init__()
        self.path=path
        self.ncfile=nc.Dataset(path)
        self.input_lon=float(input_lon)
        self.input_lat=float(input_lat)
        self.function_type=function_type
        self.time_num=time_num  #当前时间的编号
        self.time=time  #当前时间
        self.all_time_index=all_time_index  #所有的时间
        self.speedup_index=speedup_index

    def run(self):
        # 重写线程执行的run函数
        # 触发自定义信号
        self.lon=to_np(getvar(self.ncfile, 'lon'))
        self.lat=to_np(getvar(self.ncfile, 'lat'))
        ua = getvar(self.ncfile, 'ua', self.time_num)
        ua = to_np(ua)
        va = getvar(self.ncfile,'va',self.time_num)
        va = to_np(va)
        if self.speedup_index==0:#表示没有进行加速计算
            lon_x=self.lon
            lat_x=self.lat
            ua_s=ua
            va_s=va
        else:#进行了加速计算，通过index+1等于加速的方式省略了elif
            x_y=self.search_nearest_point()
            all_line_index, all_point_index = self.lon.shape[0], self.lon.shape[1]
            self.input_line_index, self.input_point_index = int(x_y[0]), int(x_y[1])
            #计算四个参数：线条的起始点，线条上划分网格的起始点，线条参考wrf输出信息的图形结构
            line_start=int(self.input_line_index-all_line_index/(self.speedup_index+1)/2)
            line_end=int(self.input_line_index+all_line_index/(self.speedup_index+1)/2)
            point_start=int(self.input_point_index-all_point_index/(self.speedup_index+1)/2)
            point_end=int(self.input_point_index+all_point_index/(self.speedup_index+1)/2)
            #加一个判断，防止计算的起始点和终止点超出矩阵边界
            if line_start<0:
                line_start=0
            if line_end>int(self.lon.shape[0]):
                line_end=int(self.lon.shape[0])
            if point_start<0:
                point_start=0
            if point_end>int(self.lon.shape[1]):
                point_end=int(self.lon.shape[1])
            self.slice_num=[line_start,line_end,point_start,point_end]
            #对经纬度和风速进行切片
            lon_x=self.lon[line_start:line_end,point_start:point_end]
            lat_x=self.lat[line_start:line_end,point_start:point_end]
            ua_s=ua[:,line_start:line_end,point_start:point_end]
            va_s=va[:,line_start:line_end,point_start:point_end]
        self.windspeed_interpolate_ua=[]#用于储存插值点的ua风速信息
        self.windspeed_interpolate_va=[]#用于储存插值点的va风速信息
        if self.time_num==self.all_time_index-1:
            print('yes')
        else:#进行逐个高度的插值
            for i in range(ua.shape[0]):
                ua_x=ua_s[i,:,:]
                va_x=va_s[i,:,:]
                ua_Rbf=Rbf(lon_x,lat_x,ua_x,function=self.function_type)
                va_Rbf=Rbf(lon_x,lat_x,va_x,function=self.function_type)
                self.windspeed_interpolate_ua.append(float(ua_Rbf(float(self.input_lon),float(self.input_lat))))
                self.windspeed_interpolate_va.append(float(va_Rbf(float(self.input_lon),float(self.input_lat))))
                # 通过自定义信号把待显示的字符串传递给槽函数
                self.trigger.emit(float((i+1)/(ua.shape[0]+1)*100))
        #写入xlsx
        self.write_to_xlsx()
        #放出进度给processbar
        self.trigger.emit(float(100))

    def write_to_xlsx(self):
        # 读取数据
        height = getvar(self.ncfile, 'height')
        height_maskedarray = to_np(height)
        height_maskedarray_1 = height_maskedarray[:, 0,0]
        height_maskedarray_2 = height_maskedarray[:, -1,-1]
        hgt_maskedarray = self.ncfile.variables['HGT']
        hgt_maskedarray_1 = hgt_maskedarray[0, 0,0 ]
        hgt_maskedarray_2 = hgt_maskedarray[0, -1,-1 ]
        height_to_earth_float=(height_maskedarray_1 - hgt_maskedarray_1+height_maskedarray_2-hgt_maskedarray_2)/2
        height_to_earth = []
        for i in height_to_earth_float:
            height_to_earth.append(int(i))
        # 创建xlsx
        workbook = openpyxl.Workbook()
        #写入风速插值
        worksheet = workbook.create_sheet()
        worksheet.title = '风速插值'
        worksheet.cell(1,1,'高度(m)')
        worksheet.cell(1,2,'风速u(m/s)')
        worksheet.cell(1,3,'风速v(m/s)')
        worksheet.cell(1,4,'风速(m/s)')
        worksheet.cell(1,5,'风向(°))')
        windspeed_interpolate=np.sqrt(np.power(self.windspeed_interpolate_ua,2)+np.power(self.windspeed_interpolate_va,2))
        windspeed_interpolat_vector = []
        for i in range(len(windspeed_interpolate)):#由于tan只能生成-90~90°的角度，而风速角度是0~360，因此需要做处理。其中正北为0度，顺时针为正。
            if self.windspeed_interpolate_ua[i]>=0 and self.windspeed_interpolate_va[i]>0:#第三象限
                windspeed_interpolat_vector.append(
                    np.arctan(np.divide(self.windspeed_interpolate_ua[i],self.windspeed_interpolate_va[i]))*180/np.pi+180)
            elif self.windspeed_interpolate_ua[i]<0 and self.windspeed_interpolate_va[i]>0:#第二象限
                windspeed_interpolat_vector.append(
                    np.arctan(np.divide(self.windspeed_interpolate_ua[i],self.windspeed_interpolate_va[i]))*180/np.pi+180)
            elif self.windspeed_interpolate_ua[i]>0 and self.windspeed_interpolate_va[i]<0:#第四象限
                windspeed_interpolat_vector.append(
                    np.arctan(np.divide(self.windspeed_interpolate_ua[i],self.windspeed_interpolate_va[i]))*180/np.pi+360)
            elif self.windspeed_interpolate_ua[i]>0 and self.windspeed_interpolate_va[i]==0:
                windspeed_interpolat_vector.append(270)
            elif self.windspeed_interpolate_ua[i]<0 and self.windspeed_interpolate_va[i]==0:
                windspeed_interpolat_vector.append(90)
            else:
                windspeed_interpolat_vector.append(
                    np.arctan(np.divide(self.windspeed_interpolate_ua[i],self.windspeed_interpolate_va[i])) * 180/np.pi)
        for i in range(len(height_to_earth)):  #循环写入风速信息
            worksheet.cell(i + 2, 1, height_to_earth[i])
            worksheet.cell(i + 2, 2, self.windspeed_interpolate_ua[i])
            worksheet.cell(i + 2, 3, self.windspeed_interpolate_va[i])
            worksheet.cell(i + 2, 4, windspeed_interpolate[i])
            worksheet.cell(i + 2, 5, windspeed_interpolat_vector[i])
        ####################################################################################
        #写入插值信息
        worksheet2 = workbook.create_sheet()
        worksheet2.title='插值信息'
        worksheet2.cell(1,1,'文件名：')
        worksheet2.cell(1,2,self.path)
        worksheet2.cell(2,1,'加速倍数：')
        worksheet2.cell(2,2,str(self.speedup_index+1)+'x')
        worksheet2.cell(3,1,'距离插值点最近网格点索引值：')
        worksheet2.cell(3,2,str(self.input_line_index)+','+str(self.input_point_index))
        worksheet2.cell(4,1,'最近网格点的经纬度坐标：')
        worksheet2.cell(4,2,str(self.lon[self.input_line_index,self.input_point_index])+
                        ','+str(self.lat[self.input_line_index,self.input_point_index]))
        worksheet2.cell(5,1,'计算的时间（UTC+0）：')
        worksheet2.cell(5,2,self.time)
        worksheet2.cell(6,1,'切片起始点索引值：')
        worksheet2.cell(6,2,str(self.slice_num[0])+':'+str(self.slice_num[1])+','
                        +str(self.slice_num[2])+':'+str(self.slice_num[3]))
        worksheet2.cell(7,1,'风向正北为0°，顺时针为正角度。')
        #保存xlsx
        workbook.save(filename=self.path[0:-3]+"_"+"ws"+".xlsx")

    def search_nearest_point(self):
        distance = (self.lon - self.input_lon) ** 2 + (self.lat - self.input_lat) ** 2
        return np.where(distance == np.min(distance))

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())