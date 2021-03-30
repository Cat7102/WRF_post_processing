#此脚本用于批量插值变量到格点，使用scipy.interpolate.rbf函数，可以自己定义加速倍数，但是我还是不建议用这玩意
#这个是用于除了风速以外的其他参数
import netCDF4 as nc
from wrf import getvar,to_np,ALL_TIMES
from datetime import datetime,timedelta
from scipy.interpolate.rbf import Rbf
import openpyxl
import numpy as np
import time

#这个函数是用于搜索最近格点的，不建议修改
def search_nearest_point(lon,lat,point_lon,point_lat):
    distance = (lon - point_lon) ** 2 + (lat - point_lat) ** 2 #计算最近点的距离
    print("最近格点的索引："+str(np.where(distance == np.min(distance))))
    return np.where(distance == np.min(distance))

#这个函数用于获取nc文件的时间，同样不建议修改
def get_ncfile_time(ncfile):
    timelist=[]
    time = str(to_np(getvar(ncfile, 'times'))) #这个输出的时间是nc文件中最开始的时间，例如2016-07-21T00:00:00.000000000
    time = time[0:-10] #把最后一些无意义的东西筛掉
    times = getvar(ncfile, 'xtimes', timeidx=ALL_TIMES) #这个输出的是分钟的时间
    formal_datetime = datetime.strptime(time, '%Y-%m-%dT%H:%M:%S') #格式化时间
    for i in times:
        timelist.append(str(formal_datetime + timedelta(minutes=int(i)))) #逐一把分钟加到最开始的时间上，并且格式化，形成全部的utc时间
    print("nc文件中所有的时间是："+str(timelist))
    return timelist

#这个函数获取离地高度的列表，不过数值是粗略计算的，如果了解的话可以修改，否则不建议修改
def get_ncfile_height2earth_vague(ncfile):
    # 读取数据
    height = to_np(getvar(ncfile, 'height')) #获取nc文件的所有海拔高度信息
    height_start = height[:, 0, 0] #获取0,0的海拔高度
    height_end = height[:, -1, -1] #获取最后一个格点的海拔高度
    hgt = ncfile.variables['HGT'] #获取nc文件的所有地面海拔高度
    hgt_start = hgt[0, 0, 0] #获取0,0的地面海拔高度
    hgt_end = hgt[0, -1, -1] #获取最后一个格点的地面海拔高度
    #由于离地高度大都差不多，因此采用最开始和最后的点做个平均的近似手段，如果要获取详细每个格点的离地高度，需要自己另做脚本计算
    height_to_earth_float=(height_start - hgt_start+height_end-hgt_end)/2
    height_to_earth = []
    for i in height_to_earth_float:
        height_to_earth.append(int(i)) #做一个取整的算法
    return height_to_earth

#这个函数用于计算风向，由于插值的风速是自己算的，因此风向也只能自己算了，不建议修改
def calculate_wind_direction(ua_list,va_list):
    windspeed_interpolat_vector = []
    for i in range(len(ua_list)):#由于tan只能生成-90~90°的角度，而风速角度是0~360，因此需要做处理。其中正北为0度，顺时针为正。
        if ua_list[i]>=0 and va_list[i]>0:#第三象限
            windspeed_interpolat_vector.append(
                np.arctan(np.divide(ua_list[i],va_list[i]))*180/np.pi+180)
        elif ua_list[i]<0 and va_list[i]>0:#第二象限
            windspeed_interpolat_vector.append(
                np.arctan(np.divide(ua_list[i],va_list[i]))*180/np.pi+180)
        elif ua_list[i]>0 and va_list[i]<0:#第四象限
            windspeed_interpolat_vector.append(
                np.arctan(np.divide(ua_list[i],va_list[i]))*180/np.pi+360)
        elif ua_list[i]>0 and va_list[i]==0:
            windspeed_interpolat_vector.append(270)
        elif ua_list[i]<0 and va_list[i]==0:
            windspeed_interpolat_vector.append(90)
        else:
            windspeed_interpolat_vector.append(
                np.arctan(np.divide(ua_list[i],va_list[i])) * 180/np.pi)
    return windspeed_interpolat_vector

#计算沿着高度进行的插值
def interpolate_wind_inheight(ncfile,height_num,function_type,input_lon,input_lat,speedup_index):
    timelist = get_ncfile_time(ncfile)
    heightlist=get_ncfile_height2earth_vague(ncfile)
    lon = to_np(getvar(ncfile, 'lon')) #读取精度并转换为numpy.ndarray格式
    lat = to_np(getvar(ncfile, 'lat'))
    TC_interpolate = []  # 用于储存插值点的T2信息
    U10_interpolate = []  # 用于储存插值点的T2信息

    if speedup_index==1:#这个表示不进行加速计算
        lon_x,lat_x=lon,lat
        line_start = 0
        line_end = int(lon.shape[0])
        point_start = 0
        point_end = int(lon.shape[1])
        slice_num = [line_start, line_end, point_start, point_end]
    else:
        # nearest_point是距离要求格点最近的点，下面一段#内的代码不建议修改
        ##################################################################
        nearest_point = search_nearest_point(lon,lat,input_lon,input_lat)
        input_line_index, input_point_index = int(nearest_point[0]), int(nearest_point[1])
        all_line_index, all_point_index = lon.shape[0], lon.shape[1]
        # 计算四个参数：线条的起始点，线条上划分网格的起始点，线条参考wrf输出信息的图形结构
        line_start = int(input_line_index - all_line_index / (speedup_index) / 2)
        line_end = int(input_line_index + all_line_index / (speedup_index) / 2)
        point_start = int(input_point_index - all_point_index / (speedup_index) / 2)
        point_end = int(input_point_index + all_point_index / (speedup_index) / 2)
        if line_start < 0:
            line_start = 0
        if line_end > int(lon.shape[0]):
            line_end = int(lon.shape[0])
        if point_start < 0:
            point_start = 0
        if point_end > int(lon.shape[1]):
            point_end = int(lon.shape[1])
        slice_num = [line_start, line_end, point_start, point_end]
        # 对经纬度进行切片
        lon_x = to_np(getvar(ncfile, 'lon'))[line_start:line_end, point_start:point_end]
        lat_x = to_np(getvar(ncfile, 'lat'))[line_start:line_end, point_start:point_end]
        ##################################################################

    for i in range(len(timelist)):
        # 这里仅做两个示范，采用所有高度下的温度和10米高的u，展示一下不同的数据获取方式，不同的数据都有最简单的获取方式。
        # 如果有疑问可以联系我或者寻找getvar和ncfile自带的变量解决
        # 如果需要添加参数，需要这几部：
        #   1.添加读取的参数变量
        #   2.对增加的参数进行模拟函数
        #   3.新建列表并在列表中加入插值点的计算结果
        #   4.在xlsx中添加一列作为写入信息
        TC = to_np(getvar(ncfile, 'tc', i))[height_num, line_start:line_end, point_start:point_end]
        U10 = ncfile.variables["U10"][i,line_start:line_end, point_start:point_end]
        TC_Rbf = Rbf(lon_x, lat_x, TC, function=function_type)
        U10_Rbf = Rbf(lon_x, lat_x, U10, function=function_type)
        TC_interpolate.append(float(TC_Rbf(float(input_lon), float(input_lat))))
        U10_interpolate.append(float(U10_Rbf(float(input_lon), float(input_lat))))
        print("第"+str(i)+"个插值计算完毕")

    # 写入xlsx
    write_to_xlsx(timelist,TC_interpolate,U10_interpolate,
                  input_line_index,input_point_index,
                  lon,lat,heightlist[height_num],speedup_index=speedup_index,slice_num=slice_num)



#这个函数用于将数据写入xlsx，可以根据自己的需求进行自定义
def write_to_xlsx(coloum_1,p1_list,p2_list,
                  input_line_index,input_point_index,lon,lat,height,speedup_index=None,slice_num=None):
    # 创建xlsx
    workbook = openpyxl.Workbook()
    #写入风速插值
    worksheet = workbook.create_sheet()
    worksheet.title = '风速插值'
    worksheet.cell(1,1,'时间') #这里可以进行修改，第一列是什么就写什么标题
    worksheet.cell(1,2,'温度(℃)')
    worksheet.cell(1,3,'风速u10(m/s)')
    for i in range(len(coloum_1)):  #循环写入风速信息
        worksheet.cell(i + 2, 1, coloum_1[i])
        worksheet.cell(i + 2, 2, p1_list[i])
        worksheet.cell(i + 2, 3, p2_list[i])
    ####################################################################################
    #写入插值信息
    worksheet2 = workbook.create_sheet()
    worksheet2.title='插值信息'
    worksheet2.cell(2,1,'加速倍数：')
    worksheet2.cell(2,2,str(speedup_index)+'x')
    worksheet2.cell(3,1,'距离插值点最近网格点索引值：')
    worksheet2.cell(3,2,str(input_line_index)+','+str(input_point_index))
    worksheet2.cell(4,1,'最近网格点的经纬度坐标：')
    worksheet2.cell(4,2,str(lon[input_line_index,input_point_index])+
                    ','+str(lat[input_line_index,input_point_index]))
    worksheet2.cell(5,1,'计算的高度(m)：')
    worksheet2.cell(5,2,height)
    worksheet2.cell(6,1,'切片起始点索引值：')
    worksheet2.cell(6,2,str(slice_num[0])+':'+str(slice_num[1])+','
                    +str(slice_num[2])+':'+str(slice_num[3]))
    worksheet2.cell(7,1,'风向正北为0°，顺时针为正角度。')
    #保存xlsx,filename是路径和最后的文件名，可以自己进行修改，可以用默认的
    workbook.save(filename="wrf_otherparameters_intime"+".xlsx")
    print("xlsx文件写入完毕")

times=time.time()

#可以看看ncfile的时间和其序列
##################################################################
ncfile=nc.Dataset('/media/fishercat/Data,Doc,Recourse etc/wrfout_d03_2016-07-21_00_00_00.nc')
heightlist=get_ncfile_height2earth_vague(ncfile=ncfile)
for i in range(len(heightlist)):
    print("["+str(i)+"]"+str(heightlist[i]))
##################################################################

#下面是建议进行修改的地方，这里做下interpolate_wind_inheight参数解释：
#   ncfile是读取的nc文件，可以修改的是nc.Dataset函数括号内的路径，要有引号，单双皆可
#   time_num是你要读取的时间的参数，可以先将下面的代码注释后运行一下上面我用的代码，看看需要的时间序列
#   function_type是插值的函数种类，具体可以参考https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.Rbf.html
#   input_lon,input_lat是你要进行插值的点
#   speedup_index是加速倍数，1表示不加速
##################################################################

ncfile=nc.Dataset('/media/fishercat/Data,Doc,Recourse etc/wrfout_d03_2016-07-21_00_00_00.nc')
interpolate_wind_inheight(ncfile=ncfile,
                          function_type='gaussian',
                          input_lon=121,input_lat=31,
                          speedup_index=5,
                          height_num=4)

##################################################################

timee=time.time()
print('本次计算总时间为：')
print(timee-times)