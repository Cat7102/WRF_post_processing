import netCDF4 as nc
from wrf import getvar,to_np
import openpyxl

#这个函数不建议修改
def write2xlsx(data1,str1,data2,str2,data3,str3):
    # 创建xlsx
    workbook = openpyxl.Workbook()
    worksheet1 = workbook.create_sheet()
    worksheet1.title = str1
    for i in range(data1.shape[0]):
        for j in range(data1.shape[1]):
            worksheet1.cell(j+1,i+1,data1[i,j])
    worksheet2 = workbook.create_sheet()
    worksheet2.title = str2
    for i in range(data2.shape[0]):
        for j in range(data2.shape[1]):
            worksheet2.cell(j+1,i+1,data2[i,j])
    worksheet3 = workbook.create_sheet()
    worksheet3.title = str3
    for i in range(data3.shape[0]):
        for j in range(data3.shape[1]):
            worksheet3.cell(j+1,i+1,data3[i,j])
    # 保存xlsx,filename是路径和最后的文件名，可以自己进行修改，可以用默认的
    workbook.save(filename="height" + ".xlsx")
    print("写入完毕")

# 读取数据，这里修改自己的路径就行
###
ncfile=nc.Dataset("/media/fishercat/Data,Doc,Recourse etc/wrfout_d03_2016-07-21_00_00_00.nc")
###
height = to_np(getvar(ncfile, 'height')) #获取nc文件的所有海拔高度信息
print("该nc文件高度一共有(层)：")
print(height.shape[0])
hgt = ncfile.variables['HGT'][0] #获取nc文件的所有地面海拔高度
###
height_layer=0 #这里可以修改数字，数字代表高度层,下面的代码就别改了，改这一行就ok了，指定第几层就是把第几层写到xlsx里,不要注释
###
point_height=height[height_layer]

height2earth=point_height-hgt #这一行不要修改，也不要注释

write2xlsx(point_height,'第'+str(height_layer)+'层海拔高度',height2earth,'第'+str(height_layer)+'层离地高度',hgt,"地面层海拔高度")
