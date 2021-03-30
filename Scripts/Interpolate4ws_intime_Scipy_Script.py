#此脚本用于批量插值变量到格点，使用scipy.interpolate.rbf函数，可以自己定义加速倍数，但是我还是不建议用这玩意
import netCDF4 as nc
from wrf import getvar,to_np,ALL_TIMES
from datetime import datetime,timedelta
from scipy.interpolate.rbf import Rbf
import openpyxl
import numpy as np
import time

#这个函数是用于搜索最近格点的，不建议修改
def search_nearest_point(lon,lat,point_lon,point_lat):
    distance = (lon - point_lon) ** 2 + (lat - point_lat) ** 2 #计算最近点的距离
    print("最近格点的索引："+str(np.where(distance == np.min(distance))))
    return np.where(distance == np.min(distance))

#这个函数用于获取nc文件的时间，同样不建议修改
def get_ncfile_time(ncfile):
    timelist=[]
    time = str(to_np(getvar(ncfile, 'times'))) #这个输出的时间是nc文件中最开始的时间，例如2016-07-21T00:00:00.000000000
    time = time[0:-10] #把最后一些无意义的东西筛掉
    times = getvar(ncfile, 'xtimes', timeidx=ALL_TIMES) #这个输出的是分钟的时间
    formal_datetime = datetime.strptime(time, '%Y-%m-%dT%H:%M:%S') #格式化时间
    for i in times:
        timelist.append(str(formal_datetime + timedelta(minutes=int(i)))) #逐一把分钟加到最开始的时间上，并且格式化，形成全部的utc时间
    print("nc文件中所有的时间是："+str(timelist))
    return timelist

#这个函数获取离地高度的列表，不过数值是粗略计算的，如果了解的话可以修改，否则不建议修改
def get_ncfile_height2earth_vague(ncfile):
    # 读取数据
    height = to_np(getvar(ncfile, 'height')) #获取nc文件的所有海拔高度信息
    height_start = height[:, 0, 0] #获取0,0的海拔高度
    height_end = height[:, -1, -1] #获取最后一个格点的海拔高度
    hgt = ncfile.variables['HGT'] #获取nc文件的所有地面海拔高度
    hgt_start = hgt[0, 0, 0] #获取0,0的地面海拔高度
    hgt_end = hgt[0, -1, -1] #获取最后一个格点的地面海拔高度
    #由于离地高度大都差不多，因此采用最开始和最后的点做个平均的近似手段，如果要获取详细每个格点的离地高度，需要自己另做脚本计算
    height_to_earth_float=(height_start - hgt_start+height_end-hgt_end)/2
    height_to_earth = []
    for i in height_to_earth_float:
        height_to_earth.append(int(i)) #做一个取整的算法
    return height_to_earth

#这个函数用于计算风向，由于插值的风速是自己算的，因此风向也只能自己算了，不建议修改
def calculate_wind_direction(ua_list,va_list):
    windspeed_interpolat_vector = []
    for i in range(len(ua_list)):#由于tan只能生成-90~90°的角度，而风速角度是0~360，因此需要做处理。其中正北为0度，顺时针为正。
        if ua_list[i]>=0 and va_list[i]>0:#第三象限
            windspeed_interpolat_vector.append(
                np.arctan(np.divide(ua_list[i],va_list[i]))*180/np.pi+180)
        elif ua_list[i]<0 and va_list[i]>0:#第二象限
            windspeed_interpolat_vector.append(
                np.arctan(np.divide(ua_list[i],va_list[i]))*180/np.pi+180)
        elif ua_list[i]>0 and va_list[i]<0:#第四象限
            windspeed_interpolat_vector.append(
                np.arctan(np.divide(ua_list[i],va_list[i]))*180/np.pi+360)
        elif ua_list[i]>0 and va_list[i]==0:
            windspeed_interpolat_vector.append(270)
        elif ua_list[i]<0 and va_list[i]==0:
            windspeed_interpolat_vector.append(90)
        else:
            windspeed_interpolat_vector.append(
                np.arctan(np.divide(ua_list[i],va_list[i])) * 180/np.pi)
    return windspeed_interpolat_vector

#计算沿着高度进行的插值
def interpolate_wind_inheight(ncfile,height_num,function_type,input_lon,input_lat,speedup_index):
    timelist = get_ncfile_time(ncfile)
    heightlist=get_ncfile_height2earth_vague(ncfile)
    lon = to_np(getvar(ncfile, 'lon')) #读取精度并转换为numpy.ndarray格式
    lat = to_np(getvar(ncfile, 'lat'))
    windspeed_interpolate_ua = []  # 用于储存插值点的ua风速信息
    windspeed_interpolate_va = []  # 用于储存插值点的va风速信息
    windspeed_interpolate_wa = []  # 用于储存插值点的wa风速信息

    if speedup_index==1:#这个表示不进行加速计算
        lon_x,lat_x=lon,lat
        line_start = 0
        line_end = int(lon.shape[0])
        point_start = 0
        point_end = int(lon.shape[1])
        slice_num = [line_start, line_end, point_start, point_end]
    else:
        # nearest_point是距离要求格点最近的点，下面一段#内的代码不建议修改
        ##################################################################
        nearest_point = search_nearest_point(lon,lat,input_lon,input_lat)
        input_line_index, input_point_index = int(nearest_point[0]), int(nearest_point[1])
        all_line_index, all_point_index = lon.shape[0], lon.shape[1]
        # 计算四个参数：线条的起始点，线条上划分网格的起始点，线条参考wrf输出信息的图形结构
        line_start = int(input_line_index - all_line_index / (speedup_index) / 2)
        line_end = int(input_line_index + all_line_index / (speedup_index) / 2)
        point_start = int(input_point_index - all_point_index / (speedup_index) / 2)
        point_end = int(input_point_index + all_point_index / (speedup_index) / 2)
        if line_start < 0:
            line_start = 0
        if line_end > int(lon.shape[0]):
            line_end = int(lon.shape[0])
        if point_start < 0:
            point_start = 0
        if point_end > int(lon.shape[1]):
            point_end = int(lon.shape[1])
        slice_num = [line_start, line_end, point_start, point_end]
        # 对经纬度和风速进行切片
        lon_x = to_np(getvar(ncfile, 'lon'))[line_start:line_end, point_start:point_end]
        lat_x = to_np(getvar(ncfile, 'lat'))[line_start:line_end, point_start:point_end]
        ##################################################################

    for i in range(len(timelist)):
        #读取数据
        ua_x = to_np(getvar(ncfile, 'ua', i))[height_num, line_start:line_end, point_start:point_end]
        va_x = to_np(getvar(ncfile, 'va', i))[height_num, line_start:line_end, point_start:point_end]
        wa_x = to_np(getvar(ncfile, 'wa', i))[height_num, line_start:line_end, point_start:point_end]
        #调用Rbf函数进行模拟
        ua_Rbf = Rbf(lon_x, lat_x, ua_x, function=function_type)
        va_Rbf = Rbf(lon_x, lat_x, va_x, function=function_type)
        wa_Rbf = Rbf(lon_x, lat_x, wa_x, function=function_type)
        #向列表中添加插值点的计算结果
        windspeed_interpolate_ua.append(float(ua_Rbf(float(input_lon), float(input_lat))))
        windspeed_interpolate_va.append(float(va_Rbf(float(input_lon), float(input_lat))))
        windspeed_interpolate_wa.append(float(wa_Rbf(float(input_lon), float(input_lat))))
        print("第"+str(i)+"个插值计算完毕")
    #对水平风速进行计算
    windspeed_interpolate = np.sqrt(np.power(windspeed_interpolate_ua, 2) + np.power(windspeed_interpolate_va, 2))
    winddirection_interpolate = calculate_wind_direction(windspeed_interpolate_ua,windspeed_interpolate_va)
    # 添加最近点网格的参数
    nearest_point_ua = []
    nearest_point_va = []
    nearest_point_wa = []
    for i in range(len(timelist)):
        nearest_point_ua.append(
            to_np(getvar(ncfile, 'ua', i))[height_num, input_line_index, input_point_index])
        nearest_point_va.append(
            to_np(getvar(ncfile, 'va', i))[height_num, input_line_index, input_point_index])
        nearest_point_wa.append(
            to_np(getvar(ncfile, 'wa', i))[height_num, input_line_index, input_point_index])

    # 写入xlsx
    write_to_xlsx(timelist,windspeed_interpolate_ua,windspeed_interpolate_va,windspeed_interpolate_wa,windspeed_interpolate,
                  winddirection_interpolate,nearest_point_ua,nearest_point_va,nearest_point_wa,
                  input_line_index,input_point_index,lon,lat,heightlist[height_num],speedup_index=speedup_index,slice_num=slice_num)



#这个函数用于将数据写入xlsx，可以根据自己的需求进行自定义
def write_to_xlsx(coloum_1,ua_list,va_list,wa_list,ws_list,wdir_list,nearest_ua_list,nearest_va_list,nearest_wa_list,
                  input_line_index,input_point_index,lon,lat,height,speedup_index=None,slice_num=None):
    # 创建xlsx
    workbook = openpyxl.Workbook()
    #写入风速插值
    worksheet = workbook.create_sheet()
    worksheet.title = '风速插值'
    worksheet.cell(1,1,'时间') #这里可以进行修改，第一列是什么就写什么标题
    worksheet.cell(1,2,'风速u(m/s)')
    worksheet.cell(1,3,'风速v(m/s)')
    worksheet.cell(1,4,'风速w(m/s)')
    worksheet.cell(1,5,'风速(m/s)')
    worksheet.cell(1,6,'风向(°))')
    worksheet.cell(1,7,'最近网格点的风速u(m/s)')
    worksheet.cell(1,8,'最近网格点的风速v(m/s)')
    worksheet.cell(1,9,'最近网格点的风速w(m/s)')
    for i in range(len(coloum_1)):  #循环写入风速信息
        worksheet.cell(i + 2, 1, coloum_1[i])
        worksheet.cell(i + 2, 2, ua_list[i])
        worksheet.cell(i + 2, 3, va_list[i])
        worksheet.cell(i + 2, 4, wa_list[i])
        worksheet.cell(i + 2, 5, ws_list[i])
        worksheet.cell(i + 2, 6, wdir_list[i])
        worksheet.cell(i + 2, 7, nearest_ua_list[i])
        worksheet.cell(i + 2, 8, nearest_va_list[i])
        worksheet.cell(i + 2, 9, nearest_wa_list[i])
    ####################################################################################
    #写入插值信息
    worksheet2 = workbook.create_sheet()
    worksheet2.title='插值信息'
    worksheet2.cell(2,1,'加速倍数：')
    worksheet2.cell(2,2,str(speedup_index)+'x')
    worksheet2.cell(3,1,'距离插值点最近网格点索引值：')
    worksheet2.cell(3,2,str(input_line_index)+','+str(input_point_index))
    worksheet2.cell(4,1,'最近网格点的经纬度坐标：')
    worksheet2.cell(4,2,str(lon[input_line_index,input_point_index])+
                    ','+str(lat[input_line_index,input_point_index]))
    worksheet2.cell(5,1,'计算的高度(m)：')
    worksheet2.cell(5,2,height)
    worksheet2.cell(6,1,'切片起始点索引值：')
    worksheet2.cell(6,2,str(slice_num[0])+':'+str(slice_num[1])+','
                    +str(slice_num[2])+':'+str(slice_num[3]))
    worksheet2.cell(7,1,'风向正北为0°，顺时针为正角度。')
    #保存xlsx,filename是路径和最后的文件名，可以自己进行修改，可以用默认的
    workbook.save(filename="wrf_windspeed_intime"+".xlsx")
    print("xlsx文件写入完毕")

times=time.time()

#可以看看ncfile的时间和其序列
##################################################################
ncfile=nc.Dataset('/media/fishercat/Data,Doc,Recourse etc/wrfout_d03_2016-07-21_00_00_00.nc')
heightlist=get_ncfile_height2earth_vague(ncfile=ncfile)
for i in range(len(heightlist)):
    print("["+str(i)+"]"+str(heightlist[i]))
##################################################################

#下面是建议进行修改的地方，这里做下interpolate_wind_inheight参数解释：
#   ncfile是读取的nc文件，可以修改的是nc.Dataset函数括号内的路径，要有引号，单双皆可
#   time_num是你要读取的时间的参数，可以先将下面的代码注释后运行一下上面我用的代码，看看需要的时间序列
#   function_type是插值的函数种类，具体可以参考https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.Rbf.html
#   input_lon,input_lat是你要进行插值的点
#   speedup_index是加速倍数，1表示不加速
##################################################################

ncfile=nc.Dataset('/media/fishercat/Data,Doc,Recourse etc/wrfout_d03_2016-07-21_00_00_00.nc')
interpolate_wind_inheight(ncfile=ncfile,
                          height_num=4,
                          function_type='gaussian',
                          input_lon=121,input_lat=31,
                          speedup_index=5)

##################################################################

timee=time.time()
print('本次计算总时间为：')
print(timee-times)
